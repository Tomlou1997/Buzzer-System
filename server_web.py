"""
抢答软件 - Web版服务端 (FastAPI + WebSocket)
"""
import asyncio
import json
import os
import sys
import time
import uuid
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
import subprocess
from openpyxl import load_workbook

def get_local_ip() -> str:
    """获取本机局域网IP"""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

# ==================== 数据模型 ====================
class Player:
    def __init__(self, name: str, websocket: WebSocket):
        self.name = name
        self.ws: WebSocket = websocket
        self.score = 0
        self.banned = False
        self.connected = True
        self.extend_remaining = 1  # 每场比赛可求助啦啦队次数
        self.ranked = False        # 是否已获得排名
        self.rank = 0              # 排名名次

class Question:
    def __init__(self, q_type: str, question: str, options: list, answer: str, points: int = 10):
        self.q_type = q_type
        self.question = question
        self.options = options  # ["A. xxx", "B. xxx", ...]
        self.answer = answer
        self.points = points
        self.used = False

# ==================== 游戏状态 ====================
class GameState:
    def __init__(self):
        self.players: dict[str, Player] = {}  # name -> Player
        self.questions: list[Question] = []
        self.active_bank_name = ""
        self.question_banks: dict[str, list[dict]] = {}
        
        self.game_started = False    # 是否已开始比赛（进入比赛模式）
        self.round_active = False    # 本轮抢答是否进行中
        self.game_over = False       # 比赛是否已结束
        
        self.current_question_index = -1
        self.round_num = 0
        self.first_buzzer: Optional[str] = None
        self.timer_remaining = 0
        self.timer_task: Optional[asyncio.Task] = None
        
        self.ranked_players: list[tuple[str, int, int]] = []  # [(name, score, rank)]
        
        # 设置
        self.correct_points = 2
        self.wrong_points = 1
        self.answer_timeout = 15
        self.win_score = 20
        self.win_rank_count = 3
        self.allow_repeat = False
        self.extend_max = 1
        self.extend_seconds = 15
        self.buzz_timeout = 15  # 抢答等待秒数
        self.show_restart_btn = False
        self.show_end_btn = True
        self.auto_judge = True
        self.server_port = 8888
        self.server_host = get_local_ip()  # 自动获取本机局域网IP
        self.start_time = time.time()  # 服务启动时间戳
        
        # 当前抢答者正在答题
        self.current_answerer: Optional[str] = None
        self.last_answer = ""
        
        # 答案选项映射（抢答后才设置）
        self.current_options = []
        self.current_correct = ""
        
    def reset_round(self):
        self.round_active = False
        self.first_buzzer = None
        self.current_answerer = None
        self.last_answer = ""
        self.timer_remaining = 0
        if self.timer_task:
            self.timer_task.cancel()
            self.timer_task = None


game = GameState()

# ==================== FastAPI ====================
app = FastAPI(title="抢答系统")

# 静态文件
os.makedirs("templates", exist_ok=True)

@app.get("/")
async def root():
    """根路径重定向"""
    return HTMLResponse("""
    <html><head><meta charset="utf-8"><title>抢答系统</title>
    <style>body{background:#0f0f1a;color:#fff;display:flex;align-items:center;justify-content:center;min-height:100vh;font-family:sans-serif;flex-direction:column;gap:20px;}
    a{color:#0078d4;text-decoration:none;padding:12px 30px;border:1px solid #333;border-radius:10px;display:inline-block;font-size:16px;}
    a:hover{border-color:#0078d4;background:#1a1a2e;}
    .links{display:flex;gap:20px;}</style></head>
    <body>
    <h1>🎯 抢答系统</h1>
    <div class="links">
      <a href="/server">👑 管理端</a>
      <a href="/client">🔔 客户端</a>
    </div>
    </body></html>
    """)

LOCKED_PAGE = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>管理端已锁定</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:-apple-system,BlinkMacSystemFont,"Microsoft YaHei",sans-serif;background:#0f0f1a;color:#fff;display:flex;align-items:center;justify-content:center;min-height:100vh;}
.box{text-align:center;max-width:400px;padding:40px;}
.box .icon{font-size:72px;margin-bottom:16px;}
.box h1{font-size:24px;color:#f44336;margin-bottom:12px;}
.box p{color:#888;font-size:16px;line-height:1.8;margin-bottom:24px;}
.box .btn{padding:12px 32px;background:#252538;border:1px solid #333;border-radius:8px;color:#aaa;font-size:15px;cursor:pointer;text-decoration:none;display:inline-block;}
.box .btn:hover{background:#333;color:#fff;}
@keyframes pulse{0%,100%{opacity:.6;}50%{opacity:1;}}
.box .dot{display:inline-block;width:8px;height:8px;background:#f44336;border-radius:50%;margin-right:6px;animation:pulse 1.5s infinite;}
</style>
</head>
<body>
<div class="box">
  <div class="icon">🔒</div>
  <h1>⚠️ 已有管理端连接</h1>
  <p>当前已有管理员在线操作中，<br>请稍后再试或联系管理员。</p>
  <a class="btn" href="/server"><span class="dot"></span>刷新重试</a>
</div>
</body>
</html>
"""

@app.get("/server")
async def get_admin():
    """管理端页面"""
    # 已有管理端在线时返回锁定页
    active = await get_active_admin_count()
    if active > 0:
        return HTMLResponse(LOCKED_PAGE)
    
    html_path = os.path.join("templates", "index.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            content = f.read()
            # 替换 IP 占位符
            content = content.replace("__SERVER_HOST__", game.server_host)
            return HTMLResponse(content)
    return HTMLResponse("<h1>管理端页面未找到</h1><p>请确保 templates/index.html 存在</p>")

@app.get("/client")
async def get_client():
    """客户端页面"""
    html_path = os.path.join("templates", "client.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())
    return HTMLResponse("<h1>客户端页面未找到</h1><p>请确保 templates/client.html 存在</p>")

# ==================== WebSocket 管理端 ====================
admin_connections: set[WebSocket] = set()

async def get_active_admin_count() -> int:
    """返回当前活跃的管理端数量"""
    dead = set()
    for w in admin_connections:
        try:
            await asyncio.wait_for(w.send_text(json.dumps({"type": "ping"})), timeout=1)
        except:
            dead.add(w)
    admin_connections.difference_update(dead)
    return len(admin_connections)

@app.websocket("/ws/admin")
async def admin_websocket(websocket: WebSocket):
    await websocket.accept()
    
    # 清理已断开的连接
    active = await get_active_admin_count()
    
    if active > 0:
        await websocket.send_text(json.dumps({"type": "error", "msg": "已有管理端连接"}))
        await websocket.close()
        return
    
    admin_connections.add(websocket)
    try:
        # 发送初始状态
        await send_admin_state(websocket)
        
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)
            await handle_admin_msg(msg, websocket)
    except WebSocketDisconnect:
        pass
    finally:
        admin_connections.discard(websocket)

async def send_admin_state(ws: Optional[WebSocket] = None):
    """给管理端发送完整状态"""
    state = {
        "type": "admin_state",
        "players": {n: {"name": n, "score": p.score, "banned": p.banned, 
                        "connected": p.connected,
                        "ranked": p.ranked, "rank": p.rank}
                    for n, p in game.players.items()},
        "game_started": game.game_started,
        "round_active": game.round_active,
        "game_over": game.game_over,
        "round_num": game.round_num,
        "current_question_index": game.current_question_index,
        "total_questions": len(game.questions),
        "used_count": sum(1 for q in game.questions if q.used),
        "timer_remaining": game.timer_remaining,
        "current_answerer": game.current_answerer,
        "last_answer": game.last_answer,
        "first_buzzer": game.first_buzzer,
        "ranked_players": [{"name": r[0], "score": r[1], "rank": r[2]} for r in game.ranked_players],
        "question_banks": list(game.question_banks.keys()),
        "active_bank_name": game.active_bank_name,
        "settings": {
            "correct_points": game.correct_points,
            "wrong_points": game.wrong_points,
            "answer_timeout": game.answer_timeout,
            "win_score": game.win_score,
            "win_rank_count": game.win_rank_count,
            "allow_repeat": game.allow_repeat,
            "extend_max": game.extend_max,
            "extend_seconds": game.extend_seconds,
            "auto_judge": game.auto_judge,
            "show_restart_btn": game.show_restart_btn,
            "show_end_btn": game.show_end_btn,
        },
        "uptime": int(time.time() - game.start_time),  # 运行秒数
    }
    # 当前题目
    if 0 <= game.current_question_index < len(game.questions):
        q = game.questions[game.current_question_index]
        state["current_question"] = {
            "type": q.q_type,
            "question": q.question,
            "options": q.options,
            "answer": q.answer,
            "points": q.points,
        }
    
    msg = json.dumps(state, ensure_ascii=False)
    if ws:
        await ws.send_text(msg)
    else:
        for w in admin_connections:
            try:
                await w.send_text(msg)
            except:
                pass

async def broadcast_to_admin(msg: dict):
    """广播给所有管理端"""
    text = json.dumps(msg, ensure_ascii=False)
    for w in admin_connections:
        try:
            await w.send_text(text)
        except:
            pass

async def broadcast_to_players(msg: dict, exclude: list = None):
    """广播给所有选手"""
    text = json.dumps(msg, ensure_ascii=False)
    for p in game.players.values():
        if exclude and p.name in exclude:
            continue
        if p.connected:
            try:
                await p.ws.send_text(text)
            except:
                p.connected = False

async def send_to_player(name: str, msg: dict):
    """给指定选手发消息"""
    if name in game.players and game.players[name].connected:
        try:
            await game.players[name].ws.send_text(json.dumps(msg, ensure_ascii=False))
        except:
            game.players[name].connected = False

# ==================== 管理端消息处理 ====================
async def handle_admin_msg(msg: dict, ws: WebSocket):
    msg_type = msg.get("type")
    
    if msg_type == "upload_bank":
        # 上传题库由 HTTP 接口处理
        pass
    
    elif msg_type == "activate_bank":
        name = msg.get("name", "")
        if name in game.question_banks:
            await activate_bank(name)
    
    elif msg_type == "delete_bank":
        name = msg.get("name", "")
        if name in game.question_banks:
            del game.question_banks[name]
            if game.active_bank_name == name:
                game.active_bank_name = ""
                game.questions = []
            await send_admin_state()
    
    elif msg_type == "start_game":
        """开始比赛（进入比赛模式）"""
        if not game.questions:
            await ws.send_text(json.dumps({"type": "error", "msg": "请先导入题库"}))
            return
        game.game_started = True
        game.game_over = False
        game.round_num = 0
        game.ranked_players = []
        game.current_question_index = -1  # 比赛开始时不显示题目，等管理员切题
        for q in game.questions:
            q.used = False
        # 重置选手分数和排名
        for p in game.players.values():
            p.score = 0
            p.ranked = False
            p.rank = 0
            p.extend_remaining = game.extend_max
        await broadcast_to_players({"type": "game_started", "game_name": game.active_bank_name})
        # 推送给所有选手最新状态
        for name in list(game.players.keys()):
            await send_client_state(name)
        await send_admin_state()
    
    elif msg_type == "show_question":
        """预览题目"""
        idx = msg.get("index", 0)
        if 0 <= idx < len(game.questions):
            game.current_question_index = idx
            await send_admin_state()
    
    elif msg_type == "next_question":
        if game.current_question_index < len(game.questions) - 1:
            if game.current_question_index < 0:
                game.current_question_index = 0
            else:
                game.current_question_index += 1
            await send_admin_state()
            # 如果比赛已开始，同步新题目给所有选手
            if game.game_started:
                for name in list(game.players.keys()):
                    await send_client_state(name)
    
    elif msg_type == "prev_question":
        if game.current_question_index > 0:
            game.current_question_index -= 1
            await send_admin_state()
            # 如果比赛已开始，同步新题目给所有选手
            if game.game_started:
                for name in list(game.players.keys()):
                    await send_client_state(name)
    
    elif msg_type == "start_round":
        """开始一轮抢答"""
        if not game.game_started or game.game_over:
            return
        if game.round_active:
            return
        
        idx = game.current_question_index
        if idx < 0 or idx >= len(game.questions):
            await ws.send_text(json.dumps({"type": "error", "msg": "请先选择题目"}))
            return
        
        q = game.questions[idx]
        if q.used and not game.allow_repeat:
            await ws.send_text(json.dumps({"type": "error", "msg": "该题目已被使用"}))
            return
        
        game.round_active = True
        game.round_num += 1
        game.first_buzzer = None
        game.current_answerer = None
        game.last_answer = ""
        
        q.used = True
        
        await broadcast_to_players({
            "type": "round_start",
            "round": game.round_num,
            "question": q.question,
            "q_type": q.q_type,
            "options": q.options,
        })
        # 推送客户端状态（含当前题目）
        for name in list(game.players.keys()):
            await send_client_state(name)
        await send_admin_state()
        
        # 启动抢答等待超时（5秒无人抢答则结束本轮）
        game.timer_remaining = game.buzz_timeout
        game.timer_task = asyncio.create_task(buzz_wait_timer())
    
    elif msg_type == "stop_round":
        """手动结束抢答"""
        game.reset_round()
        await broadcast_to_players({"type": "round_cleared", "msg": "本轮抢答已结束"})
        await send_admin_state()
    
    elif msg_type == "judge_correct":
        """判题：答对"""
        name = game.current_answerer
        if name and name in game.players:
            pts = game.correct_points
            p = game.players[name]
            if not p.ranked:
                p.score += pts
            await broadcast_to_players({
                "type": "result",
                "name": name,
                "result": "correct",
                "score": p.score,
                "msg": f"✅ [{name}] 答对 +{pts}分"
            })
            await check_winner(name)
            await end_round()
    
    elif msg_type == "judge_wrong":
        """判题：答错"""
        name = game.current_answerer
        if name and name in game.players:
            pts = game.wrong_points
            p = game.players[name]
            if not p.ranked:
                p.score = max(0, p.score - pts)
            await broadcast_to_players({
                "type": "result",
                "name": name,
                "result": "wrong",
                "score": p.score,
                "msg": f"❌ [{name}] 答错 -{pts}分"
            })
            await end_round()
    
    elif msg_type == "penalty":
        """违规扣分"""
        name = msg.get("name", "")
        if name in game.players:
            pts = msg.get("points", 5)
            p = game.players[name]
            if not p.ranked:
                p.score = max(0, p.score - pts)
            await broadcast_to_players({
                "type": "score_update",
                "name": name,
                "score": p.score,
                "msg": f"⚠️ [{name}] 违规 -{pts}分"
            })
            await send_admin_state()
    
    elif msg_type == "set_score":
        """设置分数"""
        name = msg.get("name", "")
        if name in game.players:
            game.players[name].score = msg.get("score", 0)
            await broadcast_to_players({
                "type": "score_update",
                "name": name,
                "score": game.players[name].score,
                "msg": f"📝 [{name}] 分数已设置为 {game.players[name].score}"
            })
            # 检查是否达到获胜积分
            if game.game_started:
                await check_winner(name)
            await send_admin_state()
    
    elif msg_type == "toggle_ban":
        """禁赛/恢复"""
        name = msg.get("name", "")
        if name in game.players:
            p = game.players[name]
            p.banned = not p.banned
            await send_to_player(name, {
                "type": "ban_status",
                "banned": p.banned
            })
            await send_admin_state()
    
    elif msg_type == "restart_game":
        """重赛"""
        game.game_over = False
        game.ranked_players = []
        for p in game.players.values():
            p.score = 0
            p.ranked = False
            p.rank = 0
            p.extend_remaining = game.extend_max
        for q in game.questions:
            q.used = False
        game.reset_round()
        game.round_num = 0
        await broadcast_to_players({"type": "restart", "msg": "🔄 比赛已重置，准备开始"})
        await send_admin_state()
    
    elif msg_type == "end_game":
        """结束比赛"""
        game.game_over = True
        game.round_active = False
        game.game_started = False
        if game.timer_task:
            game.timer_task.cancel()
            game.timer_task = None
        # 发送最终排名
        rankings = get_rankings()
        await broadcast_to_players({"type": "game_over", "rankings": rankings})
        # 断开所有选手
        await asyncio.sleep(0.5)
        for p in list(game.players.values()):
            try:
                await p.ws.send_text(json.dumps({"type": "server_closed"}))
                await p.ws.close()
            except:
                pass
        game.players.clear()
        await send_admin_state()
    
    elif msg_type == "update_settings":
        """更新设置"""
        s = msg.get("settings", {})
        game.correct_points = s.get("correct_points", game.correct_points)
        game.wrong_points = s.get("wrong_points", game.wrong_points)
        game.answer_timeout = s.get("answer_timeout", game.answer_timeout)
        game.win_score = s.get("win_score", game.win_score)
        game.win_rank_count = s.get("win_rank_count", game.win_rank_count)
        game.allow_repeat = s.get("allow_repeat", game.allow_repeat)
        game.extend_max = s.get("extend_max", game.extend_max)
        game.extend_seconds = s.get("extend_seconds", game.extend_seconds)
        game.buzz_timeout = s.get("buzz_timeout", game.buzz_timeout)
        game.show_restart_btn = s.get("show_restart_btn", game.show_restart_btn)
        game.show_end_btn = s.get("show_end_btn", game.show_end_btn)
        game.auto_judge = s.get("auto_judge", game.auto_judge)
        await send_admin_state()
    
    elif msg_type == "remove_player":
        """移除选手"""
        name = msg.get("name", "")
        if name in game.players:
            p = game.players[name]
            try:
                await p.ws.send_text(json.dumps({"type": "kicked", "msg": "你已被管理员移除"}))
                await p.ws.close()
            except:
                pass
            # 从排名列表中移除
            game.ranked_players = [r for r in game.ranked_players if r[0] != name]
            del game.players[name]
            await broadcast_to_admin({"type": "player_left", "name": name})
            await send_admin_state()
    
    elif msg_type == "reset_all_scores":
        """重置所有选手积分为零"""
        for p in game.players.values():
            p.score = 0
            p.ranked = False
            p.rank = 0
        game.ranked_players = []
        await broadcast_to_players({
            "type": "scores_reset",
            "msg": "🔄 管理员已重置所有积分"
        })
        await send_admin_state()


async def activate_bank(name: str):
    """激活题库"""
    if name not in game.question_banks:
        return
    game.active_bank_name = name
    raw = game.question_banks[name]
    game.questions = []
    for r in raw:
        q = Question(
            q_type=r.get("type", ""),
            question=r.get("question", ""),
            options=r.get("options", []),
            answer=r.get("answer", ""),
            points=r.get("points", 10)
        )
        game.questions.append(q)
    if game.questions:
        game.current_question_index = 0
    await send_admin_state()

async def end_round():
    """结束本轮（判题后调用）"""
    game.reset_round()
    await broadcast_to_players({"type": "round_cleared", "msg": "本轮结束"})
    # 推送给所有选手最新状态
    for name in list(game.players.keys()):
        await send_client_state(name)
    await send_admin_state()

async def check_winner(name: str):
    """检查是否达到获胜积分"""
    p = game.players.get(name)
    if not p or p.ranked:
        return
    if p.score < game.win_score:
        return
    
    rank = len(game.ranked_players) + 1
    game.ranked_players.append((name, p.score, rank))
    p.ranked = True
    p.rank = rank
    
    await broadcast_to_players({
        "type": "rank_locked",
        "name": name,
        "rank": rank,
        "score": p.score,
    })
    
    # 检查是否已决出所有名次
    if len(game.ranked_players) >= game.win_rank_count or len(game.ranked_players) >= len(game.players):
        game.game_over = True
        game.round_active = False
        rankings = get_rankings()
        await broadcast_to_players({
            "type": "game_over", 
            "rankings": rankings,
            "ranked_players": [{"name": r[0], "score": r[1], "rank": r[2]} for r in game.ranked_players],
        })
    
    await send_admin_state()

def get_rankings():
    """获取排名列表"""
    ranked = [(r[0], r[1], r[2]) for r in game.ranked_players]
    unranked = [(n, p.score) for n, p in game.players.items() if not p.ranked]
    unranked.sort(key=lambda x: x[1], reverse=True)
    result = []
    for r in ranked:
        result.append({"name": r[0], "score": r[1], "rank": r[2]})
    # 未排名选手排在后面
    for u in unranked:
        result.append({"name": u[0], "score": u[1], "rank": 0})
    return result


# ==================== WebSocket 客户端 ====================
@app.websocket("/ws/player")
async def client_websocket(websocket: WebSocket, name: str = ""):
    await websocket.accept()
    player_name = name
    login_error = None
    
    try:
        # 如果URL参数没有name，从第一条消息获取
        if not player_name:
            data = await asyncio.wait_for(websocket.receive_text(), timeout=10)
            msg = json.loads(data)
            if msg.get("type") == "login":
                player_name = msg.get("name", "").strip()
        
        if not player_name:
            await websocket.send_text(json.dumps({"type": "error", "msg": "请输入选手名称"}))
            await websocket.close()
            player_name = None
            return
        
        # 检查重名：在线拒绝，离线允许重连
        if player_name in game.players:
            if game.players[player_name].connected:
                await websocket.send_text(json.dumps({"type": "error", "msg": "该名称已被使用"}))
                await websocket.close()
                player_name = None
                return
            # 离线，重用旧数据
            game.players[player_name].ws = websocket
            game.players[player_name].connected = True
            player = game.players[player_name]
        else:
            player = Player(player_name, websocket)
            game.players[player_name] = player
        
        # 发送登录成功
        await send_to_player(player_name, {
            "type": "login_ok",
            "name": player_name,
            "score": player.score,
        })
        
        # 发送初始客户端状态
        await send_client_state(player_name)
        
        await broadcast_to_admin({"type": "player_joined", "name": player_name})
        await send_admin_state()
        
        # 消息循环
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)
            msg_type = msg.get("type")
            
            if msg_type == "buzz":
                await handle_buzz(player_name)
            
            elif msg_type == "answer":
                await handle_answer(player_name, msg.get("answer", ""))
            
            elif msg_type == "extend":
                await handle_extend(player_name)
            
            elif msg_type == "cheer":
                await handle_extend(player_name)  # cheer 同 extend
            
            elif msg_type == "pong":
                pass
    
    except asyncio.TimeoutError:
        pass
    except WebSocketDisconnect:
        pass
    except Exception as e:
        print(f"客户端异常 [{player_name}]: {e}")
    finally:
        if player_name and player_name in game.players:
            game.players[player_name].connected = False
            await broadcast_to_admin({"type": "player_left", "name": player_name})
            await send_admin_state()


async def send_client_state(name: str):
    """给单个选手发送当前状态"""
    p = game.players.get(name)
    if not p:
        return
    
    idx = game.current_question_index
    q = game.questions[idx] if 0 <= idx < len(game.questions) else None
    
    state = {
        "type": "client_state",
        "my_score": p.score,
        "my_rank": p.rank if p.ranked else 0,
        "my_ranked": p.ranked,
        "game_started": game.game_started,
        "round_active": game.round_active,
        "game_over": game.game_over,
        "can_buzz": game.round_active and game.first_buzzer is None and not p.banned and not p.ranked,
        "can_cheer": (game.current_answerer == name and p.extend_remaining > 0),
        "cheer_remaining": p.extend_remaining,
        "round_num": game.round_num,
        "players": {n: {"name": n, "score": pl.score, "banned": pl.banned}
                    for n, pl in game.players.items()},
        "ranked_players": [{"name": r[0], "score": r[1], "rank": r[2]} for r in game.ranked_players],
    }
    
    if q:
        state["current_question"] = {
            "type": q.q_type,
            "question": q.question,
            "options": q.options,
            "answer": q.answer,
        }
    
    if game.round_active:
        state["first_buzzer"] = game.first_buzzer
        state["timer_remaining"] = game.timer_remaining
    
    await send_to_player(name, state)


async def handle_buzz(name: str):
    """处理抢答"""
    if not game.round_active or game.game_over:
        await send_to_player(name, {"type": "buzz_fail", "msg": "本轮未开始抢答"})
        return
    
    if name not in game.players:
        return
    
    p = game.players[name]
    if p.banned:
        await send_to_player(name, {"type": "buzz_fail", "msg": "你已被禁赛"})
        return
    if p.ranked:
        await send_to_player(name, {"type": "buzz_fail", "msg": "你已获得排名"})
        return
    
    if game.first_buzzer is not None:
        await send_to_player(name, {"type": "buzz_fail", "msg": f"😅 [{game.first_buzzer}] 抢先一步！"})
        return
    
    # 抢答成功
    game.first_buzzer = name
    game.current_answerer = name
    
    # 取消抢答等待超时
    if game.timer_task:
        game.timer_task.cancel()
        game.timer_task = None
    
    idx = game.current_question_index
    q = game.questions[idx] if 0 <= idx < len(game.questions) else None
    
    # 通知抢答者成功
    await send_to_player(name, {
        "type": "buzz_result",
        "success": True,
        "timeout": game.answer_timeout,
        "extend_remaining": p.extend_remaining,
        "extend_seconds": game.extend_seconds,
        "options": q.options if q else [],
    })
    
    # 通知其他选手被抢了
    for pname in list(game.players.keys()):
        if pname != name:
            await send_to_player(pname, {
                "type": "buzz_result",
                "success": False,
                "winner": name,
            })
    
    # 启动倒计时
    game.timer_remaining = game.answer_timeout
    game.timer_task = asyncio.create_task(answer_timer(name))
    
    await send_client_state(name)
    await send_admin_state()


async def handle_answer(name: str, answer: str):
    """处理选手提交答案"""
    if game.current_answerer != name:
        return
    
    game.last_answer = answer
    
    # 自动判题
    idx = game.current_question_index
    correct = ""
    options = []
    if 0 <= idx < len(game.questions):
        q = game.questions[idx]
        correct = q.answer
        options = q.options or []
    
    # 将选项索引转换为选项文本进行比对
    display_answer = answer
    if options and answer:
        # 把 "0,1,2,3" 转成 "A,B,C,D" 或选项文本
        try:
            indices = [int(i.strip()) for i in answer.split(',') if i.strip().isdigit()]
            option_texts = [options[i] for i in indices if 0 <= i < len(options)]
            # 从选项文本中提取字母前缀（如 "A. xxx" -> "A"）
            letters = []
            for t in option_texts:
                t = t.strip()
                if t and (t[0].isalpha() and (len(t) == 1 or t[1] in '. )）')):
                    letters.append(t[0].upper())
            if letters:
                display_answer = ''.join(letters)
            else:
                display_answer = ','.join(option_texts)
        except (ValueError, IndexError):
            display_answer = answer
    
    is_correct = (display_answer.upper() == correct.upper()) if display_answer else False
    is_timeout = not answer  # 提前定义，后面多个地方使用
    
    if game.auto_judge:
        p = game.players[name]

        # 通知管理端答题记录
        points_change = 0
        if is_correct:
            points_change = game.correct_points
        elif is_timeout:
            points_change = -game.wrong_points
        elif not is_correct:
            points_change = -game.wrong_points
        
        # 超时时用特殊文字
        admin_answer = "答题超时" if is_timeout else display_answer
        
        await broadcast_to_admin({
            "type": "answer_received",
            "name": name,
            "answer": admin_answer,
            "correct": correct,
            "is_correct": is_correct,
            "points": points_change,
        })

        if is_correct:
            if not p.ranked:
                p.score += game.correct_points
            # 通知抢答者
            await send_to_player(name, {
                "type": "answer_result",
                "correct": True,
                "points": game.correct_points,
                "new_score": p.score,
                "answer": display_answer,
                "correct_answer": correct,
            })
            # 通知其他选手
            await broadcast_to_players({
                "type": "result",
                "name": name,
                "result": "correct",
                "score": p.score,
                "msg": f"✅ [{name}] 答对 +{game.correct_points}分",
                "answer": display_answer,
                "correct": correct,
            }, exclude=[name])
            await check_winner(name)
        else:
            # 区分超时和答错（is_timeout 已在前面定义）
            if not p.ranked:
                if not is_timeout:
                    p.score = p.score - game.wrong_points
            # 通知抢答者
            if is_timeout:
                if not p.ranked:
                    p.score = p.score - game.wrong_points
                await send_to_player(name, {
                    "type": "answer_result",
                    "correct": False,
                    "points": -game.wrong_points,
                    "new_score": p.score,
                    "msg": f"答题超时 -{game.wrong_points}分",
                    "answer": "答题超时",
                    "correct_answer": correct,
                })
            else:
                await send_to_player(name, {
                    "type": "answer_result",
                    "correct": False,
                    "points": -game.wrong_points,
                    "new_score": p.score,
                    "msg": f"回答错误 -{game.wrong_points}分",
                    "answer": display_answer,
                    "correct_answer": correct,
                })
            # 通知其他选手
            await broadcast_to_players({
                "type": "result",
                "name": name,
                "result": "wrong",
                "score": p.score,
                "msg": f"❌ [{name}] 答错 -{game.wrong_points}分",
                "answer": display_answer,
                "correct": correct,
            }, exclude=[name])
        
        # 停止计时器
        if game.timer_task:
            game.timer_task.cancel()
            game.timer_task = None
        
        await end_round()
    else:
        # 手动判题
        await broadcast_to_admin({
            "type": "answer_received",
            "name": name,
            "answer": answer,
            "correct": correct,
        })
        await send_admin_state()


async def handle_extend(name: str):
    """处理求助啦啦队"""
    if name not in game.players:
        return
    p = game.players[name]
    
    if p.extend_remaining <= 0:
        await send_to_player(name, {"type": "extend_result", "success": False, "msg": "啦啦队次数已用完"})
        return
    
    if game.current_answerer != name:
        await send_to_player(name, {"type": "extend_result", "success": False, "msg": "当前不在答题中"})
        return
    
    p.extend_remaining -= 1
    game.timer_remaining += game.extend_seconds
    
    # 通知管理端
    await broadcast_to_admin({
        "type": "cheer_used",
        "name": name,
        "added": game.extend_seconds,
        "remaining": game.timer_remaining,
    })
    
    await broadcast_to_players({
        "type": "extend_broadcast",
        "name": name,
        "remaining": game.timer_remaining,
    })
    
    await send_to_player(name, {
        "type": "extend_result",
        "success": True,
        "msg": f"🎉 啦啦队增加了{game.extend_seconds}秒，剩余{p.extend_remaining}次",
        "remaining": p.extend_remaining,
        "time_remaining": game.timer_remaining,
    })
    
    await send_admin_state()


async def answer_timer(name: str):
    """答题倒计时"""
    try:
        while game.timer_remaining > 0:
            await asyncio.sleep(1)
            game.timer_remaining -= 1
            
            await send_to_player(name, {
                "type": "timer_tick",
                "remaining": game.timer_remaining
            })
            await send_admin_state()
        
        # 超时
        if game.current_answerer == name:
            p = game.players.get(name)
            if p and not p.ranked:
                p.score = p.score - game.wrong_points
            
            idx = game.current_question_index
            correct = game.questions[idx].answer if 0 <= idx < len(game.questions) else ""

            # 通知管理端
            await broadcast_to_admin({
                "type": "answer_received",
                "name": name,
                "answer": "答题超时",
                "correct": correct,
                "is_correct": False,
                "points": -game.wrong_points,
            })
            
            # 通知答题者
            await send_to_player(name, {
                "type": "answer_result",
                "correct": False,
                "points": -game.wrong_points,
                "new_score": p.score if p else 0,
                "msg": f"答题超时 -{game.wrong_points}分",
                "correct_answer": correct,
            })
            
            # 通知其他选手
            await broadcast_to_players({
                "type": "result",
                "name": name,
                "result": "timeout",
                "score": p.score if p else 0,
                "msg": f"⏰ [{name}] 答题超时 -{game.wrong_points}分",
                "correct": correct,
            }, exclude=[name])
            await end_round()
    
    except asyncio.CancelledError:
        pass


async def buzz_wait_timer():
    """抢答等待超时 - 无人抢答则结束本轮"""
    try:
        while game.timer_remaining > 0 and game.first_buzzer is None:
            await asyncio.sleep(1)
            game.timer_remaining -= 1
            await send_admin_state()
        
        if game.first_buzzer is None and game.round_active:
            # 无人抢答
            await broadcast_to_players({"type": "buzz_too_late"})
            await send_admin_state()
            game.reset_round()
            await broadcast_to_players({"type": "round_cleared", "msg": "无人抢答，本轮结束"})
            await send_admin_state()
    
    except asyncio.CancelledError:
        pass


# ==================== 题库上传 ====================
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@app.post("/api/restart")
async def restart_server():
    """重启服务"""
    await broadcast_to_admin({"type": "restarting", "msg": "服务正在重启..."})
    await broadcast_to_players({"type": "restarting", "msg": "服务正在重启，请稍后重新连接..."})
    await asyncio.sleep(0.5)
    python = sys.executable
    script = os.path.abspath(__file__)
    subprocess.Popen([python, script], cwd=os.path.dirname(script))
    os._exit(0)

@app.post("/api/stop")
async def stop_server():
    """关闭服务"""
    await broadcast_to_admin({"type": "restarting", "msg": "服务已关闭"})
    await asyncio.sleep(0.3)
    os._exit(0)

@app.post("/api/upload")
async def upload_bank(file: UploadFile = File(...)):
    """上传xlsx题库"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"error": "请上传 .xlsx 文件"}
    
    content = await file.read()
    save_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(save_path, "wb") as f:
        f.write(content)
    
    # 解析xlsx
    questions = parse_xlsx(save_path)
    if not questions:
        return {"error": "题库为空或格式不正确"}
    
    bank_name = file.filename.replace('.xlsx', '').replace('.xls', '')
    game.question_banks[bank_name] = questions
    
    # 自动激活
    await activate_bank(bank_name)
    
    return {"success": True, "name": bank_name, "count": len(questions)}


def parse_xlsx(path: str) -> list[dict]:
    """解析xlsx题库
    列: 1-ID, 2-出题人, 3-类别, 4-来源, 5-题型, 6-题目, 7-选项, 8-分值, 9-答案, 10-组别, 11-状态
    """
    questions = []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        q_type = str(row[4]).strip() if len(row) > 4 and row[4] else ""       # 列5: 题型
        question = str(row[5]).strip() if len(row) > 5 and row[5] else ""     # 列6: 题目
        options_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ""  # 列7: 选项
        points = int(row[7]) if len(row) > 7 and row[7] else 10               # 列8: 分值
        answer = str(row[8]).strip() if len(row) > 8 and row[8] else ""       # 列9: 答案
        
        if not question:
            continue
        
        # 解析选项（换行分割）
        options = []
        for opt in options_raw.split("\n"):
            opt = opt.strip()
            if opt:
                options.append(opt)
        # 单行选项用 | 或 / 分割
        if not options and options_raw:
            import re
            options = re.split(r'[|/]', options_raw)
            options = [o.strip() for o in options if o.strip()]
        
        questions.append({
            "type": q_type,
            "question": question,
            "options": options,
            "answer": answer,
            "points": points,
        })
    
    wb.close()
    return questions


# ==================== 启动 ====================
if __name__ == "__main__":
    PORT = 8888
    print(f"🚀 抢答系统已启动")
    print(f"   管理端: http://localhost:{PORT}")
    print(f"   客户端: http://<你的IP>:{PORT}")
    print(f"   按 Ctrl+C 停止")
    uvicorn.run(app, host="0.0.0.0", port=PORT)
